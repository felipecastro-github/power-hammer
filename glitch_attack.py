"""
Glitch attack principal script

file glitch_attack.py
Author: Felipe Castro
Date: 

THIS SOFTWARE IS PROVIDED BY OnChip-UIS 

Project: Glitch attack
Module: Principal script 

How to use:
-Edit the Variables and run as follow
>>python3 glitch_attack.py
"""
###### Variables ##############
pb			=	46	  		## To set freq -> f=(pb + 4) [MHz]
freq		=	(4+pb)*1e6	## Frequency
frac		=	11 			## To set the transistor in resistive divider Vout_glitch = VDD/(1+frac)
durationTot	=	100 		## Max glitch duration in clock cycles, Max enabled 254
repeats		=	10	 		## Max iterations in each frequency for each execution
repeat_num	= 	10			## Number of repetition of the same frequency .Example. 1 -> repetitions from 1 to repeats

###############################

from pyftdi.gpio import GpioAsyncController 	## https://eblot.github.io/pyftdi/gpio.html
import os										## To execute the terminal
from openpyxl import Workbook 					## To save excel
import subprocess,shlex							## To execute python modules
import time										## To calculete the execute times

gpio = GpioAsyncController()					## ftdi gpio config
gpio.configure('ftdi://ftdi:232h:FT07AF79/1')	## ftdi device config


def main():
	########## Data export ##########
	###### for the excel data #######
	wb 			=	Workbook()
	ruta		=	str(int(freq/1e6))+'M_'+str(frac)+'_'+str(repeat_num)+'.xlsx' ## File name. Ex. 10M_0_1.xlsx
	hoja 		=	wb.active
	hoja.title 	=	"Glitch measures"
	col_f 		=	2
	col			=	6
	hoja.cell(column=col_f, row = 1, value= "Freq.") 					## Column title
	hoja.cell(column=col_f+1, row = 1, value= "Duration [CC]")			## Column title
	hoja.cell(column=col_f+2, row = 1, value= "Energy stb")				## Column title
	hoja.cell(column=col_f+3, row = 1, value=  "Energy glitch")			## Column title
	hoja.cell(column=col_f+4, row = 1, value=  "Energy signal")			## Column title
	hoja.cell(column=col_f+5, row = 1, value=  "Differences in mem?")	## Column title
	wb.save(filename = ruta) 											## save the data in excel
	
	########## Inicialization ########
	subprocess.call(["python","resetEdgeD.py",str(freq)])					## To inicialize edge detector 
	subprocess.call(["python","config_with_esp32.py",str(pb),str(frac)])	## To inicialize the clock and divider

	########### Methodology ##########
	
	row = 3		## The first row in excel workbook
	for i in range(1+((repeat_num-1)*10),repeats+((repeat_num-1)*10)+1):
		duration = 1 				## In clock cycles
		col=6						## First column to save the resets data 
		while (duration < durationTot):				
			print('cycles:'+str(duration) + ' Repeat:'+str(i))
			pins=0
			doGiltch(duration,freq) ## Do glitch
			time.sleep(0.1)			## Wait to check the gpios
			pins = gpio.read() 		## read ftdi gpios
			#print("pins ->" + str(pins))
		
			if (((pins>>4)&1) & ((pins>>5)&1) & ~((pins>>6)&1)): 	## Only Turpial_GPIO 0 and Turpial_GPIO 1 are on -> Normal operation
				duration+=1
				print("Doing glitch...")
				#### start_time = time.time()
			elif ((pins>>6)&1): 									## One Turpial_GPIO from 2 to 7 is on
				print("Salto") 
				en=get_energy_dump_mem(duration,i) ## To get energy signals and dump the turpial memory
				if (int(en[3])):
					hoja.cell(column=col_f+5, row = row, value="Yes")
				else:
					hoja.cell(column=col_f+5, row = row, value="No")

				hoja.cell(column=col_f, row = row, value= freq)		## Put frequency in the first column
				hoja.cell(column=col_f+1, row = row, value= duration)## Put the glitch duration in the second column
				hoja.cell(column=col_f+2, row = row, value= en[0])	## Put the glitch electrical energy in 3th column 
				hoja.cell(column=col_f+3, row = row, value= en[1])	## Put the standby electrical energy in 4th column
				hoja.cell(column=col_f+4, row = row, value= en[2])	## Put the glitch signal energy in the 5th column
				wb.save(filename = ruta) 							## save the data in excel
				program()											## Program turpial
				row+=1
				#### print("Elapsed time: "+str(time.time()-start_time)+" seconds")
				break
			else:												  	## When reset or other cases 
				#hoja.cell(column=col_re, row = fila-1, value= "Reset") 
				print("Programing...")
				hoja.cell(column=col, row = row, value= duration)	## Put the reset glitch duration 
				wb.save(filename = ruta) ## save the data in excel	## Save the data in excel
				col+=1												## Increment the column counter for more resets
				program()											## Program turpial
				#TODO. Verify the programing
	wb.close()

####### Functions #########
def doGiltch(duration,f):
	subprocess.call(["python","glitch.py",str(49),str(-3.0),str(duration),str(f)])	## To do glitch. In enable mode don't import the width 

def program():
	os.system("sh program.sh")								## Program Turpial. program.sh contein ./mRISCVprog -v -h ./tests/glitch_attack.dat 															## -> https://github.com/onchipuis/mriscvprog
															## 
	subprocess.call(["python","resetEdgeD.py",str(freq)])	## Reset the edge detector in chipwhisperer fpga

def get_energy_dump_mem(cycles,rep):
	os.chdir('Energy')																					## cd ./Energy
	os.system('python3 get_energy_dump_mem.py '+str(cycles)+' '+str(freq)+' '+str(frac)+' '+str(rep))	## Calculate energy and dump turpial memory
	file=open('energy_dump_results.txt','r')															## Open file to save energy results
	data=file.read().split(',')																			## Read the energy results
	file.close()																						## Close file of save energy results
	os.system('rm energy_dump_results.txt')																## rm energy_dump_results.txt
	os.chdir('..')																						## cd ..
	return([float(i) for i in data])																	## Return the electical energy in the
																										### glitch and standby time and energy 
																										### energy of voltage(VDD) signal.

try:
	main()
except KeyboardInterrupt:	## To enable the keyboard interrupt
	os._exit()

